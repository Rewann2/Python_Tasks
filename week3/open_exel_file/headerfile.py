import openpyxl

workbook=openpyxl.Workbook()
sheet1= workbook.create_sheet(title="sheet1.xlsx")

file=open("headerfile.text","r")
value=file.read()
startvalue="IDX00"
i=1
for  i in file:
    sheet1.cell(row=i, column=1).value = i
    sheet1.cell(row=i, column=2).value = startvalue+str(i)
    i=i+1
	
workbook.save("sheet1.xlsx")